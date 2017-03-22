#! python3
# blankRowInserter.py - insert some blank row

import openpyxl
import sys

firstNum = int(sys.argv[1])
secNum = int(sys.argv[2])
thirdExc = sys.argv[3]

wbSrc = openpyxl.load_workbook(thirdExc)
sheetSrc = wbSrc.get_active_sheet()

wbDes = openpyxl.Workbook()
wbDes.create_sheet('Sheet1')
sheetDes = wbDes.get_active_sheet()

for i in range(1, firstNum):
    for j in range(1, sheetSrc.max_column):
        sheetDes.cell(row=i, column=j).value = sheetSrc.cell(row=i, column=j).value

for i in range(firstNum+secNum, sheetSrc.max_row):
    for j in range(1, sheetSrc.max_column):
        sheetDes.cell(row=i, column=j).value = sheetSrc.cell(row=i, column=j).value

wbDes.save('blank.xlsx')