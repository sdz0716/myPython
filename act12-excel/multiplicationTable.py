#! python3
# multiplicationTable.py - N*N multiplicationTable

import openpyxl
from openpyxl.styles import Font
import sys

num = int(sys.argv[1])

wb = openpyxl.Workbook()
wb.create_sheet('Sheet1')
sheet = wb.get_active_sheet()

for i in range(2, num+2):
    sheet.cell(row=1, column=i).font = Font(bold=True)
    sheet.cell(row=i, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=i).value = i-1     #cell(int，int)赋值必须带value，否则报错
    sheet.cell(row=i, column=1).value = i-1
    for j in range(2, num+2):
        sheet.cell(row=i, column=j).value = '=%s*%s' % (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=1).value)
        sheet.cell(row=j, column=i).value = '=%s*%s' % (sheet.cell(row=j, column=1).value, sheet.cell(row=1, column=i).value)
wb.save('test1.xlsx')