#! python3
# flip.py - row and column turn over

import openpyxl

wb = openpyxl.load_workbook('testflip.xlsx')
sheet = wb.get_active_sheet()

wbNew = openpyxl.Workbook()
wbNew.create_sheet('Sheet1')
sheetNew = wbNew.get_active_sheet()

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        sheetNew.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value

wbNew.save('flip.xlsx')