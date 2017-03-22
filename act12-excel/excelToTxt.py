#! python3
# excelToTxt.py - excel to txt

import openpyxl

wb = openpyxl.load_workbook('txtToExcel.xlsx')
sheet = wb.get_active_sheet()

file = open('fileExcel.txt', 'w')
for i in range(1, sheet.max_row+1):
    file.write(sheet.cell(row=i, column=1).value)

file.close()