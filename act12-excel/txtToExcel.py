#! python3
# txtToExcel.py - txt to excel

import openpyxl

file = open('file.txt', 'w')
file.write('''aaaaaaaaaaaaaaaaaa
bbbbbbbbbbbbbbbbbb
cccccccccccccccccc
dddddddddddddddddd
''')
file.close()

wb = openpyxl.Workbook()
wb.create_sheet('Sheet1')
sheet = wb.get_active_sheet()

file = open('file.txt')
i = 1
for line in file.readlines():
    sheet.cell(row=i, column=1).value = line
    i += 1
wb.save('txtToExcel.xlsx')
file.close()