#! python3
# excelToCsv.py - change excel to csv

import openpyxl
import csv
import os

os.chdir('C:\\Users\\daize\\Desktop\\pythontest\\Automate\\excel')

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue
    nameFirst = excelFile[:-5]
    wb = openpyxl.load_workbook(excelFile)
    for excelSheet in wb.get_sheet_names():
        excelFormatSheet = wb.get_sheet_by_name(excelSheet)
        csvFile = open(nameFirst+excelSheet+'.csv', 'w', newline='')
        csvWriter = csv.writer(csvFile)
        for row in range(1, excelFormatSheet.max_row+1):
            excelData = []
            for column in range(1, excelFormatSheet.max_column+1):
                excelData.append(excelFormatSheet.cell(row=row, column=column).value)
                # print(excelData)
            csvWriter.writerow(excelData)
        csvFile.close()