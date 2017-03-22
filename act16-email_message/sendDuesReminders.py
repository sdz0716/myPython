#! python3
# sendDuesReminders.py - sends emails based on payment status in spreadsheet

import smtplib
import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\daize\\Desktop\\pythontest\\Automate\\duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()      #TLS加密的连接需要此代码进行加密，SSL不需要
smtpObj.login('account', 'passwd')

for name, email in unpaidMembers:
    body = "Suject: %s dues unpaid.\nDear %s...not paid dues for %s....Thanks" % (latestMonth, name, latestMonth)
    print('Sending email to %s' % email)
    sendmailStatus = smtpObj.sendmail('my account', email, body)        #所有收件人都发送成功，会返回空


    if sendmailStatus != {}:
        print('problem send email to %s:%s' % (email, sendmailStatus))
smtpObj.quit()